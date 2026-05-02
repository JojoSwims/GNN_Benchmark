from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "four_day_schedule_template.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "4-Day Schedule"

title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
time_font = Font(name="Calibri", size=10, bold=True)
cell_font = Font(name="Calibri", size=10)

title_fill = PatternFill("solid", fgColor="1F4E78")
header_fill = PatternFill("solid", fgColor="2E75B6")
time_fill = PatternFill("solid", fgColor="DDEBF7")
alt_fill = PatternFill("solid", fgColor="F2F2F2")

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.merge_cells("A1:E1")
ws["A1"] = "Four-Day Schedule"
ws["A1"].font = title_font
ws["A1"].fill = title_fill
ws["A1"].alignment = center
ws.row_dimensions[1].height = 32

ws["A2"] = "Week of:"
ws["A2"].font = Font(bold=True)
ws["A2"].alignment = Alignment(horizontal="right", vertical="center")
ws.merge_cells("B2:E2")
ws["B2"].alignment = left
ws["B2"].border = border

headers = ["Time", "Day 1", "Day 2", "Day 3", "Day 4"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border
ws.row_dimensions[4].height = 24

start_hour = 7
end_hour = 21
row = 5
for i, hour in enumerate(range(start_hour, end_hour + 1)):
    suffix = "AM" if hour < 12 else "PM"
    display = hour if hour <= 12 else hour - 12
    if hour == 0:
        display = 12
    time_str = f"{display}:00 {suffix}"

    tc = ws.cell(row=row, column=1, value=time_str)
    tc.font = time_font
    tc.fill = time_fill
    tc.alignment = center
    tc.border = border

    fill = alt_fill if i % 2 == 1 else None
    for col in range(2, 6):
        c = ws.cell(row=row, column=col)
        c.font = cell_font
        c.alignment = left
        c.border = border
        if fill is not None:
            c.fill = fill
    ws.row_dimensions[row].height = 28
    row += 1

notes_row = row + 1
ws.cell(row=notes_row, column=1, value="Notes:").font = Font(bold=True)
ws.merge_cells(start_row=notes_row, start_column=2, end_row=notes_row + 3, end_column=5)
notes_cell = ws.cell(row=notes_row, column=2)
notes_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
notes_cell.border = border
for r in range(notes_row, notes_row + 4):
    ws.row_dimensions[r].height = 22

ws.column_dimensions["A"].width = 14
for col in range(2, 6):
    ws.column_dimensions[get_column_letter(col)].width = 26

ws.freeze_panes = "B5"
ws.print_options.horizontalCentered = True
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True

wb.save(OUTPUT)
print(f"Wrote {OUTPUT}")
